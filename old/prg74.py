# pip install openpyxl
# pip install docxtpl
# ternary if operator
# если True условие False
from openpyxl.reader.excel import load_workbook
from docxtpl import DocxTemplate

string = ''
context = {}

doc = DocxTemplate('template.docx')
wb = load_workbook(filename='personal.xlsx')
sheet = wb['Сотрудники']  # wb.active
rows = sheet.iter_rows(min_row=2, values_only=True)

for row in rows:
    num, fio, gender = row
    f_name = f'invitation_{num}.docx'
    context = {
        'dear': 'Уважаемый' if gender == 'М' else 'Уважаемая',
        'fio': fio,
        'number': num
    }
    doc.render(context)
    doc.save(f_name)

